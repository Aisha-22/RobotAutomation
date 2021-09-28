# Import Package
import openpyxl

# Load Workbook

# Creating workbook Object
wk = openpyxl.load_workbook("C:\\Users\\user\\Documents\\MyRobotLearning\\TestData.xlsx")

# Print Sheet names
print(wk.sheetnames)

# Print Active sheet
print("Active sheet is -: " + wk.active.title)

# Create Object of Any sheet on which you want to work on
sh = wk['Sheet1']
print(sh.title)

# Read one Cell Data
print(sh['A3'].value) # Pick the value of the 'A3' Cell
print(sh['B4'].value)

# Create object of the Cell
print("*** Create object of the Cell ***")
c1 = sh.cell(3,1)
print(c1.value)

# Keyword Argument
print("*** Keyword Argument ***")
c2 = sh.cell(column=1, row=3)
print(c2.value)
print(c2.row)
print(c2.column)

# Read all Rows & Cells Data
# Find Rows having the data
print("***Find Rows having the data***")
rows = sh.max_row
columns = sh.max_column
print("Total Rows are -: " + str(rows))
print("Total columns are -: " + str(columns))

# Pick all Data from Excel
print("*** Pick all Data from Excel***")
for i in range(1,rows+1):
    for j in range(1,columns+1):
        c = sh.cell(i,j)
        print(c.value)


# Another approach
print("***Another approach ***")
for r in sh['A1': 'C4']:
    for c in r:
        print(c.value)